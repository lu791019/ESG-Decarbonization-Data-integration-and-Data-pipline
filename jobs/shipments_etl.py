import base64
import json
import os
from datetime import date
from datetime import datetime
from datetime import datetime as dt
from datetime import timedelta
from io import StringIO

import numpy as np
import openpyxl
import pandas as pd
from sqlalchemy import *

from app.celery import app
from models import engine
from services.mail_service import send_fail_mail


def get_stage():
    return os.environ['FLASK_ENV'] if 'FLASK_ENV' in os.environ else 'development'


def clean_excel(excel_name, sheet_name, usecols):

    df = pd.read_excel(str(excel_name), sheet_name=str(
        sheet_name), usecols=str(usecols), skiprows=4, header=None)
    df.columns = ['site', 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    df = df.melt(id_vars=['site'], var_name='month', value_name='amount')

    return df


@app.task(name='shipment-upload')
def upload_shipment(file):
    try:
        connect_string = engine.get_connect_string()
        db = create_engine(connect_string, echo=True)

        binary_data = base64.b64decode(file)

        with open("shipments.xlsx", "wb") as f:
            f.write(binary_data)

        # Load the Excel file
        workbook = openpyxl.load_workbook('shipments.xlsx')

        # Assuming you want to read data from the first sheet
        sheet = workbook.active

        # Get the value of the fourth cell (row=1, column=4)

        version = sheet.cell(row=2, column=2).value
        pic = sheet.cell(row=1, column=2).value

        # Close the workbook
        workbook.close()

        #     data = request.json
        #     file_path = data.get('file')

        df1 = clean_excel('shipments.xlsx', '三年計畫_出貨量', "A:M")
        df1['year'] = int(version)

        df2 = clean_excel('shipments.xlsx', '三年計畫_出貨量', "O:AA")
        df2['year'] = int(version)+1

        df3 = clean_excel('shipments.xlsx', '三年計畫_出貨量', "AC:AO")
        df3['year'] = int(version)+2

        df4 = clean_excel('shipments.xlsx', '三年計畫_出貨量', "AQ:BC")
        df4['year'] = int(version)+3

        df = df1.append(df2).append(df3).append(df4)

        df['amount'].fillna(0, inplace=True)
        df = df[df['site'] != '總計']

        # Load the Excel file
        workbook = openpyxl.load_workbook('shipments.xlsx')

        # Assuming you want to read data from the first sheet
        sheet = workbook.active

        # Get the value of the fourth cell (row=1, column=4)

        ver = sheet.cell(row=2, column=2).value
        person = sheet.cell(row=1, column=2).value

        # Close the workbook
        workbook.close()

        df['version'] = ver
        df['pic'] = person

        df['last_update_time'] = dt.strptime(
            dt.now().strftime("%Y-%m-%d %H:%M:%S"), "%Y-%m-%d %H:%M:%S")

        year = df['year']
        month = df['month']
        site = df['site']
        version = df['version']

        if len(month) == 0 or len(site) == 0:

            pass

        elif df.size == 0:

            pass

        else:

            delete_query = f"""DELETE FROM app.decarb_est_shipments WHERE year IN {tuple(year)} AND month IN {tuple(month)} AND site IN {tuple(site)} AND version IN {tuple(version)}"""

            conn = db.connect()
            conn.execute(delete_query)

            df.to_sql('decarb_est_shipments', db, index=False,
                      if_exists='append', schema='app', chunksize=10000)
            conn.close()

    except Exception as inst:
        send_fail_mail('Task', inst)
